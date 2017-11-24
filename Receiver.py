import socket, _thread, re, os, xlrd, csv
from openpyxl import Workbook, load_workbook

class receiveData():

    def receiver(self):
        programlist = []
        IP = ''
        PORT = 5000
        colnum = 1

        self.sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        self.sock.bind((IP, PORT))
        print("Awaiting to receive connection...")

        while True:
            #Buffer size of data
            data, self.addr = self.sock.recvfrom(64024)
            programlist.append('\n')
            programlist.append(data.decode())

            splitelement = re.compile('\n').split
            programlist = [part for listElement in programlist for part in splitelement(listElement) if part]

            filename = ('Hostnames')
            hostname = str(programlist[0])
            print(hostname + " has connected. Receiving list of installed programs...")

            # appending list data to cells at end of workbook
            maxrow = self.ws.max_row
            self.ws.cell(row=maxrow+1, column=colnum).value = str(programlist[0])
            colnum += 1
            self.ws.cell(row=maxrow + 1, column=colnum).value = str(programlist[1])
            colnum += 1
            self.ws.cell(row=maxrow+1, column=colnum).value = str(programlist[2])
            colnum += 1
            self.ws.cell(row=maxrow + 1, column=colnum).value = str(programlist[3])
            colnum += 1
            self.ws.cell(row=maxrow + 1, column=colnum).value = str(programlist[4])
            colnum += 1
            for i in programlist[5:-1]:
                self.ws.cell(row=maxrow+1, column=colnum).value = (i)
                maxrow += 1
            colnum = 1

            self.destFile.save(self.xlsxDatabase)
            self.destFile.close()

            #writing only received hostnames to txt file for comparison with AD
            with open(filename+".txt", "a", encoding="utf-8") as plist:
                plist.write(''.join(map(str, programlist[0])))
                plist.write('\n')

            #deleting and cleaning out objects within programList
            del programlist[:]

            full_path = os.path.realpath(self.xlsxDatabase)
            filePath = (os.path.dirname(full_path))
            print("File saved as "+self.xlsxDatabase+". File located in "+filePath)
            print('\n')
            print("Awaiting to receive connection...")

    def checkWorkbook(self):
        #Loading CSV Workbook
        self.xlsxDatabase = "InventoryDatabase.xlsx"
        if os.path.exists('InventoryDatabase.xlsx') == True:
            # destFile = Workbook()
            self.destFile = load_workbook(self.xlsxDatabase)
        else:
            self.destFile = Workbook()
            self.destFile.save(self.xlsxDatabase)

        worksheet = self.destFile.active
        worksheet.title = "Inventory"
        self.ws = self.destFile.active
        # Checking if titles exist at top of workbook
        if self.ws.cell(row=1, column=1).value != "Date and time received":
            #Adding column names into database
            self.ws.cell(row=1, column=1).value = "Date and time received"
            self.ws.cell(row=1, column=2).value = "Hostname"
            self.ws.cell(row=1, column=3).value = "Sophos Last Updated on"
            self.ws.cell(row=1, column=4).value = "Last Logged-on User"
            self.ws.cell(row=1, column=5).value = "Model"
            self.ws.cell(row=1, column=6).value = "Installed Software"

    def main(self):
        self.checkWorkbook()
        self.receiver()

        while True:
            self.receiver = self.sock.accept()
            _thread.start_new_thread(self.receiver,(self))
            self.sock.close()

if __name__ == "__main__":
    receive = receiveData()
    receive.main()
